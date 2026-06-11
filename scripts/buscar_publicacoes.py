#!/usr/bin/env python3
"""
Buscador de Publicações — DJEN / Comunica PJe (CNJ)
Consulta a API pública de comunicações processuais por OAB e gera:
  - planilha .xlsx (com colunas de controle preservadas entre rodadas)
  - dashboard .html (renderizado no servidor, sem dependências externas)
  - cópia datada em Versões/

Uso:
  python buscar_publicacoes.py --oab 36267/SC --oab 132301/PR --saida "/caminho/pasta"
Opções:
  --janela N      dias de histórico exibidos no arquivo atual (padrão 30)
  --urgente N     limite de dias úteis para marcar urgência/vermelho (padrão 5)
Dependências: openpyxl  (pip install openpyxl)
"""
import argparse, urllib.request, urllib.parse, urllib.error, json, re, html, datetime, os, shutil, collections, sys

API="https://comunicaapi.pje.jus.br/api/v1/comunicacao"
MARK=["ATO ORDINATÓRIO","DESPACHO/DECISÃO","DESPACHO / DECISÃO","SENTENÇA","DECISÃO MONOCRÁTICA","DECISÃO","DESPACHO","INTIMAÇÃO DA PAUTA","INTIMAÇÃO DE PAUTA","I N T I M A Ç Ã O","INTIMAÇÃO","CITAÇÃO","EDITAL"]
EDIT=["Status","Responsável","Prazo fatal (confirmado)","Providência","Data da ciência"]
HOJE=datetime.date.today()

def parse_args():
    ap=argparse.ArgumentParser()
    ap.add_argument("--oab",action="append",required=True,help="número/UF, ex.: 36267/SC (repita para várias)")
    ap.add_argument("--saida",default=".",help="pasta de saída")
    ap.add_argument("--janela",type=int,default=30)
    ap.add_argument("--urgente",type=int,default=5)
    return ap.parse_args()

def fetch(num,uf):
    """Retorna (items, total_no_acervo). Lança RuntimeError com mensagem amigável em falhas."""
    items,pg,total=[],1,0
    while True:
        url=f"{API}?numeroOab={urllib.parse.quote(num)}&ufOab={urllib.parse.quote(uf)}&pagina={pg}&itensPorPagina=100"
        req=urllib.request.Request(url,headers={"Accept":"application/json","User-Agent":"buscador-djen/1.0"})
        try:
            with urllib.request.urlopen(req,timeout=90) as r:
                d=json.load(r)
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"o servidor do CNJ respondeu com erro HTTP {e.code}. Tente novamente em alguns minutos.")
        except urllib.error.URLError as e:
            raise RuntimeError(f"nao foi possivel conectar a API do CNJ ({e.reason}). Verifique a internet e tente de novo.")
        except json.JSONDecodeError:
            raise RuntimeError("a API do CNJ retornou uma resposta inesperada (possivel instabilidade). Tente novamente.")
        total=d.get("count",0) or 0
        b=d.get("items",[]) or []
        items+=b
        if len(items)>=total or not b: break
        pg+=1
        if pg>200: break
    return items,total

def clean(t):
    if not t: return ""
    t=re.sub(r"<br\s*/?>"," ",t,flags=re.I); t=re.sub(r"<[^>]+>"," ",t)
    t=re.sub(r"body\{[^}]*\}[;]?","",t); t=re.sub(r"#div\w+\{[^}]*\}[;]?","",t)
    return re.sub(r"\s+"," ",html.unescape(t)).strip()

def resumo(t):
    b=t; pos=-1
    for m in MARK:
        i=b.find(m)
        if i!=-1 and i+len(m)>pos: pos=i+len(m)
    if pos!=-1: b=b[pos:]
    b=re.split(r"Intimado\(s\)\s*/\s*Citado\(s\)|Publique-se|Registre-se|Intime-se\.",b)[0]
    b=re.sub(r"^\s*Destinat[áa]rio\(a\):?\s*[^.]*?\.\s*","",b)
    b=re.sub(r"\b[A-ZÀ-Ú][A-ZÀ-Úa-zà-ú\. ]+/[A-Z]{2},\s*\d{1,2}\s+de\s+\w+\s+de\s+\d{4}.*$","",b)
    b=re.sub(r"\s+"," ",b).strip(" .:-—")
    if len(b)>220: b=b[:220].rsplit(" ",1)[0]+"…"
    return b or "(ver documento oficial)"

def prox_util(d):
    while d.weekday()>=5: d+=datetime.timedelta(days=1)
    return d
def add_uteis(d,n):
    while n>0:
        d+=datetime.timedelta(days=1)
        if d.weekday()<5: n-=1
    return d
def uteis(a,b):
    n=0; d=a
    while d<b:
        d+=datetime.timedelta(days=1)
        if d.weekday()<5: n+=1
    return n
def du_signed(fim):
    if not fim: return None
    return uteis(HOJE,fim) if fim>=HOJE else -uteis(fim,HOJE)

def prazo(t,dd):
    m=re.search(r"prazo\s+de\s+(\d+)\s*(?:\([^)]*\))?\s*dias?\s*(úteis|corridos)?",t,flags=re.I)
    if m:
        dias=int(m.group(1)); tipo=m.group(2) or "úteis"
        try: d0=datetime.datetime.strptime(dd,"%d/%m/%Y").date()
        except Exception: return (f"{dias} dias {tipo} (a conferir)",None)
        ini=prox_util(prox_util(d0+datetime.timedelta(days=1))+datetime.timedelta(days=1))
        fim=add_uteis(ini,max(0,dias-1)) if tipo=="úteis" else ini+datetime.timedelta(days=max(0,dias-1))
        return (f"{dias} dias {tipo} — vence ~ {fim.strftime('%d/%m/%Y')} (a conferir)",fim)
    if re.search(r"no prazo legal",t,flags=re.I): return ("prazo legal — conferir nos autos",None)
    return ("conferir nos autos",None)

def safe_link(u):
    u=(u or "").strip()
    return u if (u[:7].lower()=="http://" or u[:8].lower()=="https://") else ""

def main():
    args=parse_args()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    oabs=[]
    for o in args.oab:
        o=o.strip().upper().replace(" ","")
        if "/" not in o: sys.exit(f"OAB inválida: {o} (use número/UF, ex.: 36267/SC)")
        num,uf=o.split("/",1); oabs.append((re.sub(r"\D","",num),uf))
    base=os.path.abspath(args.saida); os.makedirs(os.path.join(base,"Versões"),exist_ok=True)
    ATUAL=os.path.join(base,"Buscador_Publicacoes.xlsx"); DASH=os.path.join(base,"Dashboard_Publicacoes.html")
    FB="Calibri"  # fonte neutra

    prev={}
    if os.path.exists(ATUAL):
        try:
            wbp=load_workbook(ATUAL); wsp=wbp["Publicações"]; hdr=[c.value for c in wsp[1]]; ci=hdr.index("ID")
            for r in range(2,wsp.max_row+1):
                vals=[c.value for c in wsp[r]]; idv=vals[ci]
                if idv is None: continue
                prev[str(idv)]={col:(vals[hdr.index(col)] if col in hdr else None) for col in EDIT}
        except Exception as e:
            print(f"[aviso] não foi possível ler o controle anterior ({e}); colunas de controle podem ser reiniciadas.",file=sys.stderr)

    UFS={"AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG","PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"}
    rows=[]; novos=0; diag={}; avisos=[]
    for num,uf in oabs:
        if uf not in UFS:
            sys.exit(f"UF invalida em {num}/{uf}: a sigla informada nao e um estado valido (ex.: SC, SP, PR). Corrija e rode de novo.")
        if not num:
            sys.exit(f"Numero de OAB vazio para a seccional {uf}. Informe o numero, ex.: 36267/{uf}.")
        try: items,total=fetch(num,uf)
        except RuntimeError as e: sys.exit(f"Erro ao consultar a OAB {num}/{uf}: {e}")
        n0=len(rows)
        for it in items:
            t=clean(it.get("texto")); dd=it.get("datadisponibilizacao")
            try: ddate=datetime.datetime.strptime(dd,"%d/%m/%Y").date()
            except Exception: ddate=None
            pub=prox_util(ddate+datetime.timedelta(days=1)) if ddate else None
            ptxt,fim=prazo(t,dd)
            recente=ddate and (HOJE-ddate).days<=args.janela
            aberto=fim and fim>=HOJE
            if not(recente or aberto): continue
            idv=str(it.get("id")); pe=prev.get(idv,{})
            st=pe.get("Status") or "Novo"
            if not pe: novos+=1
            A=[p.get("nome","") for p in it.get("destinatarios",[]) if p.get("polo")=="A" and p.get("nome")]
            P=[p.get("nome","") for p in it.get("destinatarios",[]) if p.get("polo")=="P" and p.get("nome")]
            f=lambda l:(l[0]+(" e outros" if len(l)>1 else "")) if l else "—"
            du=du_signed(fim)
            state="sem" if du is None else ("venc" if du<0 else "r" if du<=args.urgente else "y" if du<=10 else "g")
            rows.append({"OAB":f"{num}/{uf}","Processo":it.get("numeroprocessocommascara") or it.get("numero_processo"),
                "Autor/Réu":f"Autor: {f(A)}\nRéu: {f(P)}","Tribunal/Órgão-Vara":f"{it.get('siglaTribunal','')} — {it.get('nomeOrgao','')}",
                "Classe/Tipo":f"{(it.get('nomeClasse') or '').title()} — {it.get('tipoComunicacao','')}","Resumo do teor":resumo(t),
                "Status":st,"Responsável":pe.get("Responsável") or "","Prazo fatal (confirmado)":pe.get("Prazo fatal (confirmado)") or "",
                "Providência":pe.get("Providência") or "","Data da ciência":pe.get("Data da ciência") or "",
                "Link oficial":it.get("link") or "","ID":idv,"Data disponibilização":dd,
                "Data publicação":(pub.strftime("%d/%m/%Y") if pub else ""),"Prazo estimado (a conferir)":ptxt,
                "_trib":(it.get("siglaTribunal") or ""),"_fim":fim,"_du":du,"_state":state})
        em_aberto=len(rows)-n0
        diag[f"{num}/{uf}"]={"acervo":total,"em_aberto":em_aberto}
        if total==0:
            avisos.append(f"A OAB {num}/{uf} nao retornou NENHUMA comunicacao no DJEN. Confira se o numero e a UF estao certos. Lembre: processos em segredo de justica nao aparecem na consulta publica, e pode realmente nao haver publicacoes para esta inscricao.")
        elif em_aberto==0:
            avisos.append(f"A OAB {num}/{uf} tem {total} comunicacao(oes) no acervo, mas nenhuma dentro da janela de {args.janela} dias nem com prazo em aberto. As mais antigas seguem nas Versoes; aumente --janela para ve-las.")
    ORD={"venc":0,"r":1,"y":2,"g":3,"sem":4}
    rows.sort(key=lambda r:(r["_fim"] or datetime.date.max, ORD[r["_state"]]))

    cols=["OAB","Processo","Autor/Réu","Tribunal/Órgão-Vara","Classe/Tipo","Resumo do teor","Status","Responsável","Prazo fatal (confirmado)","Providência","Data da ciência","Link oficial","ID","Data disponibilização","Data publicação","Prazo estimado (a conferir)"]
    wb=Workbook(); ws=wb.active; ws.title="Publicações"
    hf=Font(name=FB,size=11,bold=True,color="000000"); bf=Font(name=FB,size=11,color="000000")
    thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    GRAY=PatternFill("solid",fgColor="F2F2F2"); YEL=PatternFill("solid",fgColor="FFF6CC")
    SEMA={"venc":PatternFill("solid",fgColor="F2B8B5"),"r":PatternFill("solid",fgColor="F9D5D3"),"y":PatternFill("solid",fgColor="FFEAB0"),"g":PatternFill("solid",fgColor="D4EDDA"),"sem":None}
    ws.append(cols)
    for c in range(1,len(cols)+1):
        cell=ws.cell(1,c); cell.font=hf; cell.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True); cell.border=border
    for r in rows: ws.append([r.get(c,"") for c in cols])
    lc=cols.index("Link oficial")+1; edit_idx={cols.index(c)+1 for c in EDIT}
    for i,r in enumerate(rows,start=2):
        baseF=SEMA[r["_state"]] or (GRAY if i%2==0 else None)
        for c in range(1,len(cols)+1):
            cell=ws.cell(i,c); cell.alignment=Alignment(vertical="top",wrap_text=True); cell.border=border
            if baseF: cell.fill=baseF
            if c in edit_idx: cell.fill=YEL
            _u=safe_link(cell.value) if c==lc else ""
            if c==lc and _u:
                cell.hyperlink=_u; cell.value="abrir documento"; cell.font=Font(name=FB,size=11,color="0000EE",underline="single")
            else: cell.font=bf
    W={"OAB":11,"Processo":24,"Autor/Réu":26,"Tribunal/Órgão-Vara":32,"Classe/Tipo":26,"Resumo do teor":50,"Status":13,"Responsável":16,"Prazo fatal (confirmado)":17,"Providência":22,"Data da ciência":13,"Link oficial":15,"ID":12,"Data disponibilização":13,"Data publicação":13,"Prazo estimado (a conferir)":30}
    for idx,c in enumerate(cols,1): ws.column_dimensions[get_column_letter(idx)].width=W[c]
    ws.freeze_panes="B2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
    cnt=collections.Counter(r["_state"] for r in rows)
    obs=wb.create_sheet("Observações")
    oablbl=", ".join(f"{n}/{u}" for n,u in oabs)
    for i,l in enumerate([
      "Buscador de Publicações — DJEN / Comunica PJe (CNJ)",
      f"Gerado em: {HOJE.strftime('%d/%m/%Y')}  |  OABs: {oablbl}",
      f"Em aberto: {len(rows)}. Vencidos(est.): {cnt['venc']} · ≤{args.urgente}du: {cnt['r']} · 6-10du: {cnt['y']} · 11+du: {cnt['g']} · sem prazo: {cnt['sem']}.","",
      f"Semáforo por dias úteis até o vencimento estimado: vermelho ≤{args.urgente}, amarelo 6-10, verde 11+, vermelho escuro vencido, sem cor = sem prazo no teor.",
      "Ordenação: cronológica crescente (vence primeiro no topo); sem prazo ao final.",
      "Colunas de controle (amarelas) preenchidas pela controladoria e preservadas por ID: Status, Responsável, Prazo fatal (confirmado), Providência, Data da ciência.",
      "Data publicação = 1º dia útil seguinte à disponibilização (base da contagem; sempre conferir).",
      "Prazo estimado é apoio: o CNJ não devolve prazo estruturado; só é extraído quando escrito no teor. A data fatal real é a fixada pela controladoria.",
      "Versionamento: cópia datada em 'Versões/'. Cobertura: Estadual, Federal, Trabalho e Superiores. Fora: STF (facultativo) e Eleitoral/Militar (não validado). Atos por mandado/carta/edital não passam pelo diário."],1):
        obs.cell(i,1,l).font=Font(name=FB,size=11,bold=(i in(1,5)),color="000000")
    obs.column_dimensions["A"].width=120
    wb.save(ATUAL)
    shutil.copy2(ATUAL,os.path.join(base,"Versões",f"Buscador_Publicacoes_{HOJE.strftime('%Y-%m-%d')}.xlsx"))

    # ---------- DASHBOARD (server-rendered) ----------
    PILL={"venc":"#7b241c","r":"#e74c3c","y":"#e0a800","g":"#28a745","sem":"#bbb"}
    esc=lambda s: html.escape(str(s if s is not None else ""))
    por_trib=dict(collections.Counter(r["_trib"] for r in rows).most_common())
    def bars(d,color="#5b8def"):
        mx=max(d.values()) if d else 1
        return "".join(f"<div class=bar><span class=bl>{esc(k)}</span><span class=bt style='width:{max(6,int(v/mx*100))}%;background:{color}'></span><span class=bn>{v}</span></div>" for k,v in d.items()) or "<div class=muted>—</div>"
    def bars_state():
        lab={"venc":"Vencido (est.)","r":f"≤{args.urgente} dias úteis","y":"6–10 dias úteis","g":"11+ dias úteis","sem":"Sem prazo no teor"}
        mx=max(cnt.values()) if cnt else 1
        return "".join(f"<div class=bar><span class=bl>{lab[k]}</span><span class=bt style='width:{max(6,int(cnt.get(k,0)/mx*100))}%;background:{PILL[k]}'></span><span class=bn>{cnt.get(k,0)}</span></div>" for k in ["venc","r","y","g","sem"])
    def trrow(r):
        venc=f"<span class=pill style='background:{PILL[r['_state']]}'>{r['_fim'].strftime('%d/%m/%Y')}</span>" if r["_fim"] else "<span class=muted>—</span>"
        du="—" if r["_du"] is None else (f"venc {abs(r['_du'])}" if r["_du"]<0 else r["_du"])
        _lk=safe_link(r["Link oficial"]); lk=f"<a href='{esc(_lk)}' target=_blank rel=noopener>abrir</a>" if _lk else ""
        ar=esc(r["Autor/Réu"]).replace("\n"," · ")
        txt=esc((str(r["Processo"])+" "+r["Autor/Réu"]+" "+r["Resumo do teor"]).lower())
        return (f"<tr class={r['_state']} data-oab=\"{esc(r['OAB'])}\" data-trib=\"{esc(r['_trib'])}\" data-status=\"{esc(r['Status'])}\" data-resp=\"{esc(r['Responsável'])}\" data-state=\"{r['_state']}\" data-txt=\"{txt}\">"
                f"<td>{venc}</td><td>{du}</td><td>{esc(r['Data publicação'])}</td><td>{esc(r['OAB'])}</td><td>{esc(r['Processo'])}</td>"
                f"<td>{esc(r['_trib'])}</td><td>{esc(r['Status'])}</td><td>{esc(r['Responsável'])}</td><td>{esc(r['Resumo do teor'])}</td><td>{lk}</td></tr>")
    TBODY="".join(trrow(r) for r in rows)
    oablist=sorted(set(r["OAB"] for r in rows)); tribs=sorted(set(r["_trib"] for r in rows))
    stats=sorted(set(r["Status"] for r in rows)); resps=sorted(set(r["Responsável"] for r in rows if r["Responsável"]))
    opt=lambda L:"".join(f"<option value=\"{esc(x)}\">{esc(x)}</option>" for x in L)
    HTML=f"""<!doctype html><html lang=pt-br><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Painel de Publicações — DJEN</title>
<style>
:root{{color-scheme:light}}*{{box-sizing:border-box}}
body{{font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;margin:0;background:#fafafa;color:#1a1a1a}}
.wrap{{max-width:1320px;margin:0 auto;padding:24px}}
h1{{font-size:22px;margin:0 0 2px}}.sub{{color:#666;font-size:13px;margin-bottom:20px}}
.kpis{{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin-bottom:22px}}
.kpi{{background:#fff;border:1px solid #ececec;border-radius:12px;padding:14px}}
.kpi .n{{font-size:26px;font-weight:700}}.kpi .l{{color:#666;font-size:11px;text-transform:uppercase;letter-spacing:.03em}}
.kpi.venc .n{{color:#7b241c}}.kpi.r .n{{color:#c0392b}}.kpi.y .n{{color:#9a7d0a}}.kpi.g .n{{color:#1e7e34}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:22px}}
.card{{background:#fff;border:1px solid #ececec;border-radius:12px;padding:16px}}
.card h3{{margin:0 0 12px;font-size:14px}}
.bar{{display:flex;align-items:center;gap:8px;margin:6px 0;font-size:12.5px}}
.bl{{width:160px;flex:none;color:#444;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.bt{{height:14px;border-radius:7px;display:inline-block}}.bn{{font-weight:600;min-width:24px;text-align:right}}
.filters{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px}}
select,input{{padding:8px 10px;border:1px solid #ddd;border-radius:8px;font-size:13px;background:#fff}}
table{{width:100%;border-collapse:collapse;background:#fff;font-size:12.5px}}
th,td{{text-align:left;padding:9px 10px;border-bottom:1px solid #f0f0f0;vertical-align:top}}
th{{background:#f7f7f7;cursor:pointer;position:sticky;top:0;z-index:1}}
tr{{border-left:5px solid transparent}}
tr.venc{{border-left-color:#7b241c;background:#fbeeec}}tr.r{{border-left-color:#e74c3c;background:#fdecea}}
tr.y{{border-left-color:#e0a800;background:#fff8e6}}tr.g{{border-left-color:#28a745;background:#f1faf3}}tr.sem{{border-left-color:#ccc}}
.pill{{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:600;color:#fff}}
.muted{{color:#999}}a{{color:#1155cc}}
@media(max-width:1000px){{.kpis{{grid-template-columns:repeat(3,1fr)}}.grid{{grid-template-columns:1fr}}}}
</style></head><body><div class=wrap>
<h1>Painel de Publicações — DJEN / Comunica PJe</h1>
<div class=sub>Gerado em {HOJE.strftime('%d/%m/%Y')} · OABs {esc(oablbl)} · retrato do dia (não recarrega ao vivo)</div>
<div class=kpis>
<div class=kpi><div class=n>{len(rows)}</div><div class=l>Em aberto</div></div>
<div class="kpi venc"><div class=n>{cnt['venc']}</div><div class=l>Vencidos (est.)</div></div>
<div class="kpi r"><div class=n>{cnt['r']}</div><div class=l>≤{args.urgente} dias úteis</div></div>
<div class="kpi y"><div class=n>{cnt['y']}</div><div class=l>6–10 dias úteis</div></div>
<div class="kpi g"><div class=n>{cnt['g']}</div><div class=l>11+ dias úteis</div></div>
<div class=kpi><div class=n>{cnt['sem']}</div><div class=l>Sem prazo / a conferir</div></div>
</div>
<div class=grid>
<div class=card><h3>Por tribunal</h3>{bars(por_trib)}</div>
<div class=card><h3>Por situação de prazo</h3>{bars_state()}</div>
</div>
<div class=card>
<div class=filters>
<input id=q placeholder="Buscar processo, parte, teor…" style=flex:1>
<select id=foab><option value="">Todas as OABs</option>{opt(oablist)}</select>
<select id=ftrib><option value="">Todos os tribunais</option>{opt(tribs)}</select>
<select id=fstatus><option value="">Todos os status</option>{opt(stats)}</select>
<select id=fresp><option value="">Todos os responsáveis</option>{opt(resps)}</select>
<select id=festado><option value="">Toda situação</option><option value=venc>Vencido (est.)</option><option value=r>≤{args.urgente} dias</option><option value=y>6–10 dias</option><option value=g>11+ dias</option><option value=sem>Sem prazo</option></select>
</div>
<div style="max-height:560px;overflow:auto"><table id=tb><thead><tr>
<th>Vencimento ▲</th><th>Dias úteis</th><th>Publicação</th><th>OAB</th><th>Processo</th><th>Tribunal</th><th>Status</th><th>Resp.</th><th>Resumo</th><th>Link</th>
</tr></thead><tbody id=body>{TBODY}</tbody></table></div>
<div id=vazio class=muted style="display:none;padding:14px">Nenhum item com os filtros atuais.</div>
</div>
<p class=muted style=font-size:11px>Semáforo por dias úteis até o vencimento estimado. Prazos são apoio (o CNJ não devolve prazo estruturado) — sempre conferir nos autos. STF (facultativo) e Eleitoral/Militar (não validado) podem não constar.</p>
</div>
<script>
const $=s=>document.querySelector(s);
function f(){{const q=$('#q').value.toLowerCase(),o=$('#foab').value,t=$('#ftrib').value,s=$('#fstatus').value,r=$('#fresp').value,e=$('#festado').value;
 let vis=0;document.querySelectorAll('#body tr').forEach(tr=>{{
  let ok=(!q||tr.dataset.txt.includes(q))&&(!o||tr.dataset.oab===o)&&(!t||tr.dataset.trib===t)&&(!s||tr.dataset.status===s)&&(!r||tr.dataset.resp===r)&&(!e||tr.dataset.state===e);
  tr.style.display=ok?'':'none';if(ok)vis++;}});
 $('#vazio').style.display=vis?'none':'block';}}
['q','foab','ftrib','fstatus','fresp','festado'].forEach(id=>$('#'+id).addEventListener('input',f));
</script></body></html>"""
    open(DASH,"w",encoding="utf-8").write(HTML)
    resultado={"ok":True,"abertas":len(rows),"novas":novos,"urgentes":cnt['r']+cnt['venc'],
        "por_estado":dict(cnt),"por_tribunal":por_trib,"por_oab":diag,"avisos":avisos,
        "arquivo":ATUAL,"dashboard":DASH}
    print(json.dumps(resultado,ensure_ascii=False))

if __name__=="__main__":
    main()
